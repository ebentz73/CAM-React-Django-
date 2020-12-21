import os
import tempfile
import openpyxl

from django.contrib.auth.models import Group
from django.db.models.signals import m2m_changed, post_save
from django.dispatch import receiver

from app.models import (
    AnalyticsSolution,
    ConstNodeData,
    FilterCategory,
    FilterOption,
    InputDataSet,
    InputDataSetInput,
    InputNodeData,
    InputPage,
    Model,
    Node,
    Scenario,
)
from app.proto_modules import (
    ConstNodeData_pb2,
    InputNodeData_pb2,
    NodeTagListBlob_pb2,
    TagFilterOption_pb2,
)
from app.utils import Sqlite

from profile.models import Role
from guardian.shortcuts import assign_perm

CAM_ROLE_PREFIX = 'CAM_ROLE=='


def get_or_create_solution_group(solution):
    groups = Group.objects.filter(name=solution.name)
    group = groups.first()
    if group is None:
        group = Group.objects.create(name=solution.name)
        assign_perm('app.view_analyticssolution', group, solution)
    return group


@receiver(post_save, sender=AnalyticsSolution)
def update_model(sender, **kwargs):
    solution = kwargs.get('instance')

    if 'tam_file' in solution.changed_fields:
        tag_roles = {}
        group = get_or_create_solution_group(solution)
        # Download tam model file and save only what we need
        f, filename = tempfile.mkstemp()
        try:
            os.write(f, solution.tam_file.read())
            os.close(f)

            # Open sqlite file
            with Sqlite(filename) as cursor:
                # Save all filters
                cursor.execute(
                    "SELECT CategoryName, FilterOptionsBlob FROM FilterCategories"
                )
                for category_name, filter_blob in cursor.fetchall():
                    category, created = FilterCategory.objects.update_or_create(
                        solution=solution, name=category_name
                    )
                    if created:
                        assign_perm('app.view_filtercategory', group, category)

                    blob_model = TagFilterOption_pb2.List_TagFilterOption()
                    blob_model.ParseFromString(filter_blob)
                    for option in blob_model.items:
                        filter_option, created = FilterOption.objects.update_or_create(
                            category=category,
                            tag=option.Tag,
                            display_name=option.DisplayName,
                        )
                        if created:
                            assign_perm('app.view_filteroption', group, filter_option)

                # Save all models
                cursor.execute("SELECT ModelId, ModelName FROM TruNavModel")
                for model_id, model_name in cursor.fetchall():
                    model, created = Model.objects.update_or_create(
                        solution=solution,
                        tam_id=model_id,
                        defaults={'name': model_name},
                    )
                    if created:
                        assign_perm('app.view_model', group, model)

                    # Save all input pages
                    cursor.execute(
                        "SELECT PageId, PageName "
                        "FROM ModelDataPage "
                        "WHERE ModelId=? AND pagetype=0",
                        (model_id,),
                    )
                    for page_id, page_name in cursor.fetchall():
                        page, created = InputPage.objects.update_or_create(
                            model=model, tam_id=page_id, defaults={'name': page_name}
                        )
                        if created:
                            assign_perm('app.view_inputpage', group, page)

                    # Save all nodes
                    cursor.execute(
                        "SELECT n.NodeId, NodeName, NodeNotes, NodeType, TagList, NodeInputData "
                        "FROM Node AS n, NodeScenarioData AS d "
                        "WHERE n.NodeId = d.NodeId "
                        "   AND n.ModelId=? "
                        "   AND NodeType IN ('inputnode', 'constnode')",
                        (model_id,),
                    )
                    for (
                        node_id,
                        node_name,
                        node_notes,
                        node_type,
                        tag_list_blob,
                        node_data_blob,
                    ) in cursor.fetchall():
                        # Get tags
                        blob_model = NodeTagListBlob_pb2.List_String()
                        blob_model.ParseFromString(tag_list_blob)
                        tag_list = list(blob_model.items)

                        # Check if node has CAM tags
                        has_tags = any(True for tag in tag_list if tag.startswith('CAM_'))

                        # Only save nodes if they are used by CAM
                        if has_tags:
                            # Create Node model
                            node, _ = Node.objects.update_or_create(
                                model=model,
                                tags=tag_list,
                                tam_id=node_id,
                                defaults={'name': node_name},
                                notes=node_notes,
                            )

                            node_data = node_data_codename = None
                            # Create InputNodeData model
                            if node_type == 'inputnode':
                                blob_model = InputNodeData_pb2.InputNodeData()
                                blob_model.ParseFromString(node_data_blob)
                                node_data = [
                                    [
                                        val.InputData.LowerBound,
                                        val.InputData.Low,
                                        val.InputData.Nominal,
                                        val.InputData.High,
                                        val.InputData.UpperBound,
                                    ]
                                    for val in blob_model.LayerData
                                ]
                                node_data, _ = InputNodeData.objects.update_or_create(
                                    node=node, default_data=node_data, is_model=True
                                )
                                node_data_codename = 'app.view_inputnodedata'

                            # Create ConstNodeData model
                            elif node_type == 'constnode':
                                blob_model = ConstNodeData_pb2.ConstNodeData()
                                blob_model.ParseFromString(node_data_blob)
                                node_data = [
                                    val.ConstData for val in blob_model.AllLayerData
                                ]
                                node_data, _ = ConstNodeData.objects.update_or_create(
                                    node=node, default_data=node_data, is_model=True
                                )
                                node_data_codename = 'app.view_constnodedata'

                            # Apply CAM role to Node
                            # only create roles for nodes we are saving
                            if node_data and node_data_codename:
                                cam_roles = (
                                    tag[len(CAM_ROLE_PREFIX):]
                                    for tag in tag_list
                                    if tag.startswith(CAM_ROLE_PREFIX)
                                )
                                for cam_role in cam_roles:
                                    node_role = tag_roles.get(cam_role)
                                    if node_role is None:
                                        # Create role since it does not exist
                                        role_name = f'{solution.name} - {cam_role}'
                                        node_role = Role.objects.create(name=role_name)
                                        tag_roles[cam_role] = node_role
                                    assign_perm('app.view_node', node_role, node)
                                    assign_perm(node_data_codename, node_role, node_data)

        finally:
            os.remove(filename)


@receiver(post_save, sender=InputDataSet)
def retrieve_spreadsheet_values(sender, **kwargs):
    ids = kwargs.get('instance')
    print(ids.input_choices())
    if 'file' in ids.changed_fields:
        wb = openpyxl.load_workbook(ids.file)

        # Input Nodes Worksheet
        input_ws = wb["Input Nodes"]
        input_ids_data = {}
        for row in input_ws.iter_rows(min_row=2):
            node_name = row[0].value
            if node_name is None:
                break
            node = Node.objects.filter(name=node_name).first()
            if node is not None:
                lowest = row[7].value
                low = row[4].value
                base = row[5].value
                high = row[6].value
                highest = row[8].value
                if node_name in input_ids_data:
                    input_ids_data[node_name]['data'].append([lowest, low, base, high, highest])
                else:
                    input_ids_data[node_name] = {}
                    input_ids_data[node_name]['data'] = [[lowest, low, base, high, highest]]
                    input_ids_data[node_name]['node'] = node
        for node in input_ids_data:
            node_data, _ = InputNodeData.objects.update_or_create(
                node=input_ids_data[node]['node'], default_data=input_ids_data[node]['data'],
                is_model=False, input_data_set=ids
            )

        # Constant Nodes Worksheet
        const_ws = wb["Constant Nodes"]
        const_ids_data = {}
        for row in const_ws.iter_rows(min_row=2):
            node_name = row[0].value
            if node_name is None:
                break
            node = Node.objects.filter(name=node_name).first()
            if node is not None:
                value = row[4].value
                if node_name in const_ids_data:
                    const_ids_data[node_name]['data'].append(value)
                else:
                    const_ids_data[node_name] = {}
                    const_ids_data[node_name]['data'] = [value]
                    const_ids_data[node_name]['node'] = node
        for node in const_ids_data:
            node_data, _ = ConstNodeData.objects.update_or_create(
                node=const_ids_data[node]['node'], default_data=const_ids_data[node]['data'],
                is_model=False, input_data_set=ids
            )


@receiver(m2m_changed, sender=InputDataSet.scenarios.through)
def input_data_set_scenario_changed(sender, **kwargs):
    action = kwargs.get('action')
    pk_set = kwargs.get('pk_set')
    ids = kwargs.get('instance')

    if action == 'pre_add':
        for scenario_id in pk_set:
            scenario = Scenario.objects.get(pk=scenario_id)
            for input_page_id in scenario.input_data_sets.values_list(
                'input_page_id', flat=True
            ):
                if ids.input_page.id == input_page_id:
                    raise Exception(
                        f"Scenario '{scenario.name}' cannot have multiple input data sets associated with Input Page '{ids.input_page.name}'."
                    )
